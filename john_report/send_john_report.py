"""
John Schreppler — Weekly LinkedIn Engagement table
====================================================
Triggered when Pablo drops a Frank LaRosa LinkedIn export. Builds a dead-simple
2-column table (Week | Engagements), keeps a rolling weekly log so the last few
week-slots accumulate across drops, then:
  1. Emails the table to John (CC Pablo) via Office365 SMTP.
  2. POSTs the table to a Power Automate flow that DMs John on Teams.

Usage:
  python send_john_report.py            # PREVIEW: build table, open it on screen, send NOTHING
  python send_john_report.py --send     # greenlight: commit the week + email John + Teams DM
  python send_john_report.py --dry-run  # print the table to console only, no file/send
  (add a path to use a specific .xlsx instead of the newest export in Downloads)

Day to day you just use the two buttons: "1 - PREVIEW.bat" then "2 - SEND TO JOHN.bat".
Nothing reaches John until you run --send.

Env (all optional; defaults match the existing newsletter setup):
  ECP_EMAIL_PASS        Office365 password for Pablo's account
  JOHN_TEAMS_FLOW_URL   Power Automate "When an HTTP request is received" URL
  JOHN_REPORT_WEEKS     How many recent weeks to show (default 4)
"""
import sys, os, re, ssl, json, glob, smtplib, urllib.request
from datetime import datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import openpyxl

try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

# ── Config ──────────────────────────────────────────────────────────────────
SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
DOWNLOADS   = os.path.join(os.path.expanduser("~"), "Downloads")
LOG_FILE    = os.path.join(SCRIPT_DIR, "engagement_log.json")
PREVIEW_FILE = os.path.join(SCRIPT_DIR, "preview.html")
EXPORT_GLOB = "AggregateAnalytics_Frank LaRosa_*.xlsx"
WEEKS_TO_SHOW = int(os.environ.get("JOHN_REPORT_WEEKS", "4"))

SMTP_SERVER = "smtp.office365.com"
SMTP_PORT   = 587
USERNAME    = "Pabloacosta@eliteconsultingpartners.com"
PASSWORD    = os.environ.get("ECP_EMAIL_PASS", "Datamagic26$")
EMAIL_TO    = ["JohnSchreppler@eliteconsultingpartners.com"]
EMAIL_CC    = ["pabloacosta@eliteconsultingpartners.com"]

TEAMS_FLOW_URL = os.environ.get("JOHN_TEAMS_FLOW_URL", "").strip()


# ── Find the export ───────────────────────────────────────────────────────────
def newest_export() -> str | None:
    files = glob.glob(os.path.join(DOWNLOADS, EXPORT_GLOB))
    if not files:
        return None
    # Prefer the export whose filename end-date is latest; fall back to mtime.
    date_pat = re.compile(r"_(\d{4}-\d{2}-\d{2})_(\d{4}-\d{2}-\d{2})\.xlsx$")
    def sort_key(p):
        m = date_pat.search(os.path.basename(p))
        if m:
            return (1, datetime.strptime(m.group(2), "%Y-%m-%d"))
        return (0, datetime.fromtimestamp(os.path.getmtime(p)))
    return max(files, key=sort_key)


# ── Parse the ENGAGEMENT sheet into Mon–Sun weekly buckets ──────────────────────
def parse_weekly_engagements(xlsx_path: str) -> dict[str, int]:
    """Return {week_start_iso: total_engagements} bucketed Mon–Sun.
    Works for any window length (a 7-day export -> 1 bucket, 28-day -> 4)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "ENGAGEMENT" not in wb.sheetnames:
        raise ValueError(f"No ENGAGEMENT sheet in {os.path.basename(xlsx_path)}")
    ws = wb["ENGAGEMENT"]

    buckets: dict[str, int] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):  # row 1 = headers
        if not row or row[0] in (None, "", "Date"):
            continue
        raw_date, _impr, eng = row[0], row[1], row[2]
        try:
            d = raw_date if isinstance(raw_date, datetime) else datetime.strptime(str(raw_date).strip(), "%m/%d/%Y")
        except ValueError:
            continue
        try:
            eng_val = int(round(float(str(eng).replace(",", "").strip())))
        except (ValueError, TypeError):
            eng_val = 0
        week_start = (d - timedelta(days=d.weekday())).date().isoformat()  # Monday
        buckets[week_start] = buckets.get(week_start, 0) + eng_val
    return buckets


# ── Rolling log ────────────────────────────────────────────────────────────────
def load_log() -> dict[str, int]:
    if os.path.exists(LOG_FILE):
        with open(LOG_FILE, encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_log(log: dict[str, int]):
    with open(LOG_FILE, "w", encoding="utf-8") as f:
        json.dump(dict(sorted(log.items())), f, indent=2)


def week_label(week_start_iso: str) -> str:
    start = datetime.fromisoformat(week_start_iso)
    end = start + timedelta(days=6)
    if start.month == end.month:
        return f"{start:%b} {start.day}–{end.day}"
    return f"{start:%b} {start.day} – {end:%b} {end.day}"


# ── Render the 2-column table ──────────────────────────────────────────────────
def build_outputs(rows: list[tuple[str, int]]) -> tuple[str, str]:
    """rows = [(week_label, engagements), ...] oldest->newest. Returns (text, html)."""
    w = max([len("Week")] + [len(r[0]) for r in rows])
    text_lines = [f"{'Week'.ljust(w)}   Engagements",
                  f"{'-'*w}   -----------"]
    for label, eng in rows:
        text_lines.append(f"{label.ljust(w)}   {eng:,}")
    text = "\n".join(text_lines)

    html_rows = "".join(
        f"<tr><td style='padding:4px 16px 4px 0'>{label}</td>"
        f"<td style='padding:4px 0;text-align:right'>{eng:,}</td></tr>"
        for label, eng in rows
    )
    html = (
        "<p>Frank LaRosa — LinkedIn engagement by week:</p>"
        "<table style='border-collapse:collapse;font-family:Arial,sans-serif;font-size:14px'>"
        "<thead><tr>"
        "<th style='text-align:left;border-bottom:1px solid #ccc;padding:4px 16px 4px 0'>Week</th>"
        "<th style='text-align:right;border-bottom:1px solid #ccc;padding:4px 0'>Engagements</th>"
        f"</tr></thead><tbody>{html_rows}</tbody></table>"
    )
    return text, html


# ── Delivery ────────────────────────────────────────────────────────────────────
def send_email(subject: str, text: str, html: str):
    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = USERNAME
    msg["To"] = ", ".join(EMAIL_TO)
    msg["Cc"] = ", ".join(EMAIL_CC)
    msg.attach(MIMEText(text, "plain", "utf-8"))
    msg.attach(MIMEText(f"<html><body>{html}</body></html>", "html", "utf-8"))
    ctx = ssl.create_default_context()
    with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
        server.ehlo(); server.starttls(context=ctx); server.ehlo()
        server.login(USERNAME, PASSWORD)
        server.sendmail(USERNAME, EMAIL_TO + EMAIL_CC, msg.as_string())
    print(f"  Email sent to {', '.join(EMAIL_TO)} (cc {', '.join(EMAIL_CC)})")

def post_to_teams(text: str, html: str, rows: list[tuple[str, int]]):
    if not TEAMS_FLOW_URL:
        print("  Teams: JOHN_TEAMS_FLOW_URL not set — skipped.")
        return
    payload = json.dumps({
        "text": text,
        "html": html,
        "rows": [{"week": r[0], "engagements": r[1]} for r in rows],
    }).encode("utf-8")
    req = urllib.request.Request(TEAMS_FLOW_URL, data=payload,
                                 headers={"Content-Type": "application/json"})
    with urllib.request.urlopen(req, timeout=30) as resp:
        print(f"  Teams flow POST → HTTP {resp.status}")


def write_and_open_preview(subject: str, html: str):
    page = (f"<html><body style='font-family:Arial,sans-serif;padding:24px'>"
            f"<p style='color:#888;font-size:12px'>PREVIEW — not sent. "
            f"To send to John, double-click \"2 - SEND TO JOHN\".</p>"
            f"<p style='font-size:13px'><b>Subject:</b> {subject}<br>"
            f"<b>To:</b> {', '.join(EMAIL_TO)} &nbsp; <b>Cc:</b> {', '.join(EMAIL_CC)}</p><hr>"
            f"{html}</body></html>")
    with open(PREVIEW_FILE, "w", encoding="utf-8") as f:
        f.write(page)
    print(f"  Preview written → {PREVIEW_FILE}")
    try:
        os.startfile(PREVIEW_FILE)  # opens in default browser (Windows)
    except Exception:
        pass


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    flags = {"--send", "--dry-run"}
    args = [a for a in sys.argv[1:] if a not in flags]
    send = "--send" in sys.argv
    dry_run = "--dry-run" in sys.argv

    export = args[0] if args else newest_export()
    if not export or not os.path.exists(export):
        print(f"ERROR: no export found (looked for {EXPORT_GLOB} in {DOWNLOADS})")
        sys.exit(1)
    print(f"Export: {os.path.basename(export)}")

    new_buckets = parse_weekly_engagements(export)
    log = load_log()
    prospective = dict(log)
    prospective.update(new_buckets)  # upsert — re-dropping a week overwrites with latest

    recent = sorted(prospective.items())[-WEEKS_TO_SHOW:]
    rows = [(week_label(ws), eng) for ws, eng in recent]
    text, html = build_outputs(rows)
    subject = f"Frank LaRosa LinkedIn — Weekly Engagement (through {week_label(recent[-1][0])})"

    print("\n" + text + "\n")

    if dry_run:
        print("DRY RUN — nothing sent, log not updated. Subject:\n  " + subject)
        return

    if not send:
        # Preview mode (default / what the watcher runs): show on screen, send nothing.
        write_and_open_preview(subject, html)
        print("PREVIEW only — log not updated, nothing sent to John.")
        print('Greenlight by double-clicking "2 - SEND TO JOHN.bat"')
        return

    # --send: commit the week to the rolling log, then deliver.
    save_log(prospective)
    send_email(subject, text, html)
    post_to_teams(text, html, rows)
    print("Done — sent to John and logged.")


if __name__ == "__main__":
    main()
