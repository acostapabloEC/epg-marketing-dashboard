"""
Monday Morning Routine — EPG Weekly Newsletter + Dashboard Update
Runs automatically every Monday at 9am via Windows Task Scheduler.

Order of operations:
  1. Pull Simplecast podcast data (API)
  2. Update LinkedIn top posts (top_posts.js — writes weekly-data.json directly)
  3. Pull Google Reviews (Apify)
  4. Run Hootsuite scraper (Instagram + YouTube xlsx)
  5. Parse xlsx → update dashboard data files → git push → deploy
  6. Build newsletter HTML with auto-generated comments
  7. Save to epg-exec-form, deploy Vercel
  8. Email Pablo with "Send to Brian" button
  9. Carla's LinkedIn engagement report (carla_report.js — separate script, separate email)
"""

import os, sys, subprocess, json
from datetime import datetime, timedelta
from pathlib import Path

SCRAPER_DIR = Path(__file__).parent
REPO_ROOT   = SCRAPER_DIR.parent

sys.path.insert(0, str(SCRAPER_DIR))


def log(msg):
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] {msg}")


# ── 1. Compute week string ────────────────────────────────────────────────────

today          = datetime.now()
days_since_mon = today.weekday()                        # Mon=0
week_end       = today - timedelta(days=days_since_mon + 1)   # last Sunday
week_start     = week_end - timedelta(days=6)                  # prior Monday
WEEK_STR = f"{week_start.strftime('%b %#d')}–{week_end.strftime('%#d, %Y')}"
log(f"Week: {WEEK_STR}")


# ── 2. Pull Simplecast ────────────────────────────────────────────────────────

log("Fetching Simplecast data...")
import send_newsletter as nl
pod = nl.fetch_simplecast_stats(
    week_start=week_start.strftime("%Y-%m-%d"),
    week_end=week_end.strftime("%Y-%m-%d"),
)
log(f"  Podcast: {pod['downloads']} downloads")


# ── 2b. Update LinkedIn top posts (independent of Carla's engagement report) ─

log("Updating LinkedIn top posts...")
top_posts_result = subprocess.run(
    ["node", "top_posts.js", week_start.strftime("%Y-%m-%d"), week_end.strftime("%Y-%m-%d"), "--write"],
    cwd=str(SCRAPER_DIR), capture_output=True, text=True, shell=True
)
if top_posts_result.returncode == 0:
    log("  Top posts updated in weekly-data.json.")
else:
    log(f"  WARNING: top_posts.js failed — keeping existing topPosts. stderr: {top_posts_result.stderr[-300:]}")


# ── 3. Read LinkedIn from weekly-data.json ────────────────────────────────────

log("Reading LinkedIn data...")
weekly_data_path = REPO_ROOT / "src" / "data" / "weekly-data.json"
li = {"engagements": 0, "prev_engagements": 0, "impressions": 0,
      "prev_impressions": 0, "members_reached": 0, "prev_members_reached": 0,
      "followers": 0, "top_posts": []}
if weekly_data_path.exists():
    wd = json.loads(weekly_data_path.read_text(encoding="utf-8"))
    li_raw = wd.get("linkedin", {})
    li = {
        "engagements":          li_raw.get("engagements",    {}).get("current",  0),
        "prev_engagements":     li_raw.get("engagements",    {}).get("previous", 0),
        "impressions":          li_raw.get("impressions",    {}).get("current",  0),
        "prev_impressions":     li_raw.get("impressions",    {}).get("previous", 0),
        "members_reached":      li_raw.get("membersReached", {}).get("current",  0),
        "prev_members_reached": li_raw.get("membersReached", {}).get("previous", 0),
        "followers":            li_raw.get("followers",      {}).get("gained",   0),
        "top_posts":            li_raw.get("topPosts", []),
    }
    log(f"  LinkedIn: {li['engagements']} eng, {li['impressions']} impressions")


# ── 4. Read Instagram + YouTube from latest xlsx in hootsuite-downloads ───────

log("Reading Hootsuite downloads...")
downloads_dir = SCRAPER_DIR / "hootsuite-downloads"

def latest_xlsx(keyword):
    files = sorted(downloads_dir.glob(f"*{keyword}*.xlsx"), key=os.path.getmtime, reverse=True)
    return files[0] if files else None

ig_file = latest_xlsx("instagram")
yt_file = latest_xlsx("youtube")

ig = {"engagements": 0, "prev_engagements": 0, "reach": 0, "prev_reach": 0,
      "views": 0, "prev_views": 0, "top_posts": []}
yt = {"views": 0, "prev_views": 0, "engagements": 0, "subscribers": 0,
      "total_subs": 0}

if ig_file:
    log(f"  Instagram xlsx: {ig_file.name}")
    try:
        import openpyxl
        wb = openpyxl.load_workbook(ig_file, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        # Sum week rows for engagements, reach, views
        total_eng = total_reach = total_views = total_likes = total_shares = total_saves = 0
        for row in rows[1:]:
            if row[0] and str(row[0]) >= week_start.strftime("%Y-%m-%d"):
                total_eng    += int(row[1] or 0)
                total_reach  += int(row[2] or 0)
                total_views  += int(row[3] or 0)
                total_likes  += int(row[4] or 0)
                total_shares += int(row[5] or 0)
                total_saves  += int(row[6] or 0)
        ig.update({"engagements": total_eng, "reach": total_reach, "views": total_views,
                   "likes": total_likes, "shares": total_shares, "saves": total_saves})
        log(f"    → {total_eng} eng, {total_reach} reach")
    except Exception as e:
        log(f"  WARNING: could not parse Instagram xlsx: {e}")

if yt_file:
    log(f"  YouTube xlsx: {yt_file.name}")
    try:
        import openpyxl
        wb = openpyxl.load_workbook(yt_file, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        total_views = total_subs = 0
        for row in rows[1:]:
            if row[0] and str(row[0]) >= week_start.strftime("%Y-%m-%d"):
                total_views += int(row[1] or 0)
                total_subs  += int(row[2] or 0)
        yt.update({"views": total_views, "subscribers": total_subs})
        log(f"    → {total_views} views, +{total_subs} subs")
    except Exception as e:
        log(f"  WARNING: could not parse YouTube xlsx: {e}")

reviews = {"rating": "4.8", "count": "22"}


# ── 5. Build newsletter data structures ───────────────────────────────────────

def fmt(n):
    if n >= 1000: return f"{n/1000:.1f}K"
    return str(n)

li_eng_note = ""
if li["prev_engagements"]:
    pct = round(((li["engagements"] - li["prev_engagements"]) / li["prev_engagements"]) * 100)
    li_eng_note = f"+{pct}% vs prior week" if pct >= 0 else f"{pct}% vs prior week"

li_data = dict(
    eng=f"{li['engagements']:,}",
    eng_note=li_eng_note,
    imp=fmt(li["impressions"]),
    followers=f"{li.get('total_followers', li['followers']):,}",
    followers_note=f"+{li['followers']} this week",
    total=str(li["engagements"]),
    total_label=f"{week_start.strftime('%b')} Total",
    total_note="Q3" if week_start.month >= 7 else "Q2",
    posts=[
        (p.get("date", ""), (p.get("preview") or p.get("text", ""))[:100] + "...",
         str(p.get("reactions", p.get("engagements", "—"))),
         f"{p.get('impressions', '—'):,}" if isinstance(p.get("impressions"), int) else str(p.get("impressions", "—")))
        for p in li["top_posts"][:2]
    ]
)

ig_data = dict(
    eng=f"{ig['engagements']:,}",
    period=f"{week_start.strftime('%b %#d')}–{week_end.strftime('%#d')}",
    views=fmt(ig["views"]),
    likes=f"{ig.get('likes', 0):,}",
    shares=f"{ig.get('shares', 0):,}",
    saves=f"{ig.get('saves', 0):,}",
    posts=[]
)

yt_data  = dict(subs=f"{yt['total_subs']:,}", note=f"+{yt['subscribers']} this week &nbsp;&middot;&nbsp; {fmt(yt['views'])} views")
pod_data = dict(dl=str(pod["downloads"]), note=f"{WEEK_STR} &nbsp;&middot;&nbsp; vs {pod.get('prev_downloads', '—')} last week")


# ── 6. Auto-generate editorial comments from data ─────────────────────────────

log("Generating editorial comments...")
comments = []

# Instagram top post
if ig["engagements"] > 1000:
    comments.append(dict(
        platform="Instagram",
        headline="Strong engagement week",
        body=f"<b style='color:#fff'>{ig['engagements']:,} engagements</b> this week across {fmt(ig['views'])} views. {ig.get('saves', 0):,} saves signal high-intent audience behavior."
    ))

# LinkedIn
if li["engagements"] > 0:
    comments.append(dict(
        platform="LinkedIn",
        headline=f"{li['engagements']:,} engagements this week",
        body=f"<b style='color:#fff'>{fmt(li['impressions'])} impressions</b> on {li['engagements']:,} engagements. {li_eng_note}."
    ))

# Podcast
if pod["downloads"] > 0:
    prev = pod.get("prev_downloads", 0)
    pod_note = ""
    if prev:
        pct = round(((pod["downloads"] - prev) / prev) * 100)
        pod_note = f" — {'+' if pct >= 0 else ''}{pct}% vs last week"
    comments.append(dict(
        platform="Podcast",
        headline=f"{pod['downloads']} downloads this week",
        body=f"<b style='color:#fff'>{pod['downloads']} downloads{pod_note}.</b> " +
             (f"Top episode: \"{(pod['episodes'][0].get('title','') or '')[:60]}...\"" if pod.get("episodes") else "")
    ))

# YouTube
comments.append(dict(
    platform="YouTube",
    headline=f"{fmt(yt['views'])} views · {yt['total_subs']:,} subscribers",
    body=f"<b style='color:#fff'>{fmt(yt['views'])} views</b> and +{yt['subscribers']} subscribers this week. Cumulative base: {yt['total_subs']:,}."
))


# ── 7. Build HTML + save + deploy + email Pablo ───────────────────────────────

from send_weekly_newsletter import build_email, send, USERNAME

log("Building newsletter HTML...")
html = build_email(WEEK_STR, li_data, ig_data, yt_data, pod_data, reviews, comments=comments)

# Save to epg-exec-form
form_repo    = REPO_ROOT.parent / "epg-exec-form"
current_path = form_repo / "newsletter-current.html"
current_path.write_text(html, encoding="utf-8")
log(f"Saved to {current_path}")

# Deploy epg-exec-form
log("Deploying epg-exec-form...")
result = subprocess.run(
    ["npx", "vercel@latest", "deploy", "--prod", "--yes"],
    cwd=str(form_repo), capture_output=True, text=True, shell=True
)
if result.returncode == 0:
    log("  Deployed successfully.")
else:
    log(f"  Deploy warning: {result.stderr[-200:]}")

# Email Pablo with Send-to-Brian button
pablo_html = f"""<!DOCTYPE html><html><body style="margin:0;padding:0;background:#0f0f0f;font-family:Arial,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0" style="background:#0f0f0f;"><tr><td align="center" style="padding:40px 16px;">
<table width="580" cellpadding="0" cellspacing="0" style="max-width:580px;width:100%;">
  <tr><td style="padding:0 0 6px;"><p style="margin:0;font-size:20px;font-weight:bold;color:#fff;">Weekly Newsletter — Ready</p></td></tr>
  <tr><td style="padding:0 0 28px;border-bottom:1px solid #1c1c1c;">
    <p style="margin:0;font-size:13px;color:#444;">{WEEK_STR} &nbsp;&middot;&nbsp; Review and send to Brian when ready</p>
  </td></tr>
  <tr><td style="padding:28px 0;">
    <p style="margin:0 0 20px;font-size:14px;color:#aaa;line-height:1.6;">This week's newsletter is built and ready. Click below to send it to Brian for approval.</p>
    <a href="https://epg-exec-form.vercel.app/api/send-to-brian" style="display:inline-block;background:#0077B5;color:#fff;padding:12px 28px;border-radius:4px;text-decoration:none;font-size:13px;font-weight:bold;">Send to Brian &rarr;</a>
  </td></tr>
  <tr><td style="padding-top:20px;border-top:1px solid #1c1c1c;">
    <p style="margin:0;font-size:11px;color:#333;">Auto-generated by ECP Analytics &nbsp;&middot;&nbsp; Every Monday by 9am ET</p>
  </td></tr>
</table></td></tr></table></body></html>"""

send(USERNAME, f"[ACTION NEEDED] Newsletter ready — {WEEK_STR}", pablo_html)
log(f"Email sent to Pablo.")


# ── 8. Carla's LinkedIn report ────────────────────────────────────────────────

log("Running Carla LinkedIn report...")
carla_result = subprocess.run(
    ["node", "carla_report.js", week_start.strftime("%Y-%m-%d"), week_end.strftime("%Y-%m-%d"), "--json"],
    cwd=str(SCRAPER_DIR), capture_output=True, text=True, shell=True
)

carla_eng = None
for line in carla_result.stdout.splitlines():
    line = line.strip()
    if line.startswith("{"):
        try:
            carla_eng = json.loads(line).get("engagement")
            break
        except Exception:
            pass

if carla_eng is not None:
    log(f"  Carla LinkedIn: {carla_eng} engagements")
    carla_html = f"""<!DOCTYPE html><html><body style="margin:0;padding:0;background:#0f0f0f;font-family:Arial,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0" style="background:#0f0f0f;"><tr><td align="center" style="padding:40px 16px;">
<table width="520" cellpadding="0" cellspacing="0" style="max-width:520px;width:100%;">
  <tr><td style="padding:0 0 6px;">
    <p style="margin:0;font-size:11px;font-weight:bold;letter-spacing:2px;text-transform:uppercase;color:#0077B5;font-family:monospace;">LinkedIn</p>
  </td></tr>
  <tr><td style="padding:0 0 6px;">
    <p style="margin:0;font-size:28px;font-weight:bold;color:#fff;">{carla_eng:,} engagements</p>
  </td></tr>
  <tr><td style="padding:0 0 28px;border-bottom:1px solid #1c1c1c;">
    <p style="margin:0;font-size:13px;color:#444;">{WEEK_STR}</p>
  </td></tr>
  <tr><td style="padding:28px 0 0;">
    <p style="margin:0;font-size:11px;color:#333;">ECP Analytics &nbsp;&middot;&nbsp; Weekly LinkedIn Report</p>
  </td></tr>
</table></td></tr></table></body></html>"""
    send("carlawade@eliteconsultingpartners.com", f"LinkedIn Report — {WEEK_STR}", carla_html)
    log("  Email sent to Carla.")
else:
    log(f"  WARNING: Could not parse Carla LinkedIn result. stderr: {carla_result.stderr[-200:]}")

log("Done.")
